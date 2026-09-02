#!/usr/bin/env python3
"""Export recent GitLab pipeline and runner details to an Excel workbook.

Required environment variables:

    export GITLAB_URL="https://gitlab.example.com"
    export GITLAB_TOKEN="glpat-..."

Example:

    python gitlab_pipeline_runner_audit.py \
      --group-id 86 \
      --days 30 \
      --excel ~/Downloads/gitlab-pipeline-audit-group-86.xlsx

The Pipeline_activity sheet contains only:

    project_id
    path_with_namespace
    last_triggered_at
    last_pipeline_status
    gitlab_runner_description
    gitlab_runner_ip_from_api

Projects without a pipeline in the lookback window are retained with blank
pipeline and runner fields. Runner IP may be blank on GitLab versions that no
longer expose it or when the token cannot read runner details.
"""

import argparse
import json
import os
import sys
import time
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode
from urllib.request import Request, urlopen


def api_get(
    base_url: str,
    token: str,
    path: str,
    params: Optional[Dict[str, Any]] = None,
    retries: int = 3,
) -> Tuple[Any, Dict[str, str]]:
    """Call one GitLab API endpoint and return JSON plus response headers."""
    url = f"{base_url.rstrip('/')}/api/v4/{path.lstrip('/')}"
    if params:
        clean = {key: value for key, value in params.items() if value is not None}
        url = f"{url}?{urlencode(clean, doseq=True)}"

    request = Request(url, headers={"PRIVATE-TOKEN": token, "Accept": "application/json"})

    for attempt in range(retries + 1):
        try:
            with urlopen(request, timeout=45) as response:
                body = response.read().decode("utf-8")
                headers = {key.lower(): value for key, value in response.headers.items()}
                return json.loads(body), headers
        except HTTPError as exc:
            if exc.code == 429 and attempt < retries:
                retry_after = int(exc.headers.get("Retry-After", "2"))
                time.sleep(max(retry_after, 1))
                continue
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"GitLab API {exc.code} for {path}: {detail[:300]}") from exc
        except URLError as exc:
            if attempt < retries:
                time.sleep(2 ** attempt)
                continue
            raise RuntimeError(f"GitLab API connection failed for {path}: {exc}") from exc

    raise RuntimeError(f"GitLab API request failed for {path}")


def paginated_get(
    base_url: str,
    token: str,
    path: str,
    params: Optional[Dict[str, Any]] = None,
) -> Iterable[Dict[str, Any]]:
    """Yield every object from a paginated GitLab list endpoint."""
    page = 1
    while True:
        query = dict(params or {})
        query.update({"page": page, "per_page": 100})
        payload, headers = api_get(base_url, token, path, query)
        if not isinstance(payload, list):
            raise RuntimeError(f"Expected a list from GitLab endpoint {path}")

        yield from payload

        next_page = headers.get("x-next-page", "")
        if next_page:
            page = int(next_page)
        elif len(payload) == 100:
            page += 1
        else:
            break


def latest_pipeline(
    base_url: str,
    token: str,
    project_id: int,
    updated_after: str,
) -> Optional[Dict[str, Any]]:
    payload, _ = api_get(
        base_url,
        token,
        f"projects/{project_id}/pipelines",
        {
            "updated_after": updated_after,
            "order_by": "updated_at",
            "sort": "desc",
            "per_page": 1,
        },
    )
    return payload[0] if payload else None


def pipeline_jobs(
    base_url: str,
    token: str,
    project_id: int,
    pipeline_id: int,
) -> List[Dict[str, Any]]:
    return list(
        paginated_get(
            base_url,
            token,
            f"projects/{project_id}/pipelines/{pipeline_id}/jobs",
            {"include_retried": "true"},
        )
    )


def runner_fields(
    base_url: str,
    token: str,
    jobs: List[Dict[str, Any]],
    cache: Dict[int, Optional[Dict[str, Any]]],
) -> Tuple[str, str]:
    """Return unique runner descriptions and API IP addresses for job runners."""
    descriptions: List[str] = []
    addresses: List[str] = []

    for job in jobs:
        runner = job.get("runner") or {}
        runner_id = runner.get("id")
        fallback_description = runner.get("description") or ""
        detail: Optional[Dict[str, Any]] = None

        if runner_id:
            numeric_id = int(runner_id)
            if numeric_id not in cache:
                try:
                    cache[numeric_id], _ = api_get(
                        base_url, token, f"runners/{numeric_id}"
                    )
                except RuntimeError:
                    # Project tokens and non-admin users may see the runner in a
                    # job but be unable to query the runner detail endpoint.
                    cache[numeric_id] = None
            detail = cache[numeric_id]

        description = (detail or {}).get("description") or fallback_description
        address = (detail or {}).get("ip_address") or ""

        if description and description not in descriptions:
            descriptions.append(str(description))
        if address and address not in addresses:
            addresses.append(str(address))

    return ", ".join(descriptions), ", ".join(addresses)


def collect_rows(
    base_url: str,
    token: str,
    group_id: int,
    days: int,
    max_projects: int,
    delay: float,
) -> Tuple[List[Dict[str, Any]], List[Tuple[int, str, str]]]:
    projects = list(
        paginated_get(
            base_url,
            token,
            f"groups/{group_id}/projects",
            {
                "include_subgroups": "true",
                "with_shared": "false",
                "simple": "true",
                "order_by": "path",
                "sort": "asc",
            },
        )
    )
    if max_projects > 0:
        projects = projects[:max_projects]

    updated_after = (
        datetime.now(timezone.utc) - timedelta(days=days)
    ).isoformat(timespec="seconds").replace("+00:00", "Z")

    rows: List[Dict[str, Any]] = []
    errors: List[Tuple[int, str, str]] = []
    runner_cache: Dict[int, Optional[Dict[str, Any]]] = {}

    for index, project in enumerate(projects, 1):
        project_id = int(project["id"])
        path = project.get("path_with_namespace") or project.get("name") or ""
        print(
            f"[{index}/{len(projects)}] {path}",
            file=sys.stderr,
            flush=True,
        )

        row: Dict[str, Any] = {
            "project_id": project_id,
            "path_with_namespace": path,
            "last_triggered_at": "",
            "last_pipeline_status": "",
            "gitlab_runner_description": "",
            "gitlab_runner_ip_from_api": "",
        }

        try:
            pipeline = latest_pipeline(
                base_url, token, project_id, updated_after
            )
            if pipeline:
                pipeline_id = int(pipeline["id"])
                row["last_triggered_at"] = (
                    pipeline.get("created_at")
                    or pipeline.get("updated_at")
                    or ""
                )
                row["last_pipeline_status"] = pipeline.get("status") or ""
                jobs = pipeline_jobs(
                    base_url, token, project_id, pipeline_id
                )
                description, address = runner_fields(
                    base_url, token, jobs, runner_cache
                )
                row["gitlab_runner_description"] = description
                row["gitlab_runner_ip_from_api"] = address
        except Exception as exc:
            errors.append((project_id, path, str(exc)))

        rows.append(row)
        if delay > 0:
            time.sleep(delay)

    rows.sort(
        key=lambda item: (
            item.get("last_triggered_at") or "",
            item.get("path_with_namespace") or "",
        ),
        reverse=True,
    )
    return rows, errors


def write_excel(
    output: Path,
    rows: List[Dict[str, Any]],
    errors: List[Tuple[int, str, str]],
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit(
            "Excel export requires openpyxl: python3 -m pip install openpyxl"
        ) from exc

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Pipeline_activity"

    headers = [
        "project_id",
        "path_with_namespace",
        "last_triggered_at",
        "last_pipeline_status",
        "gitlab_runner_description",
        "gitlab_runner_ip_from_api",
    ]
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in rows:
        sheet.append([row.get(header, "") for header in headers])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [14, 48, 25, 22, 42, 28]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    if errors:
        error_sheet = workbook.create_sheet("Errors")
        error_sheet.append(["project_id", "path_with_namespace", "error"])
        for cell in error_sheet[1]:
            cell.font = Font(bold=True)
        for error in errors:
            error_sheet.append(list(error))
        error_sheet.freeze_panes = "A2"
        error_sheet.column_dimensions["A"].width = 14
        error_sheet.column_dimensions["B"].width = 48
        error_sheet.column_dimensions["C"].width = 90

    workbook.save(output)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export GitLab project pipeline and runner details to Excel."
    )
    parser.add_argument("--group-id", type=int, required=True)
    parser.add_argument("--days", type=int, default=30)
    parser.add_argument("--excel", type=Path, required=True)
    parser.add_argument(
        "--max-projects",
        type=int,
        default=0,
        help="Limit projects for a verification run; 0 means all.",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.0,
        help="Seconds to wait after each project.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    base_url = os.environ.get("GITLAB_URL", "").strip()
    token = os.environ.get("GITLAB_TOKEN", "").strip()
    if not base_url or not token:
        print(
            "Set GITLAB_URL and GITLAB_TOKEN before running this script.",
            file=sys.stderr,
        )
        return 2
    if args.days < 1:
        print("--days must be at least 1.", file=sys.stderr)
        return 2

    rows, errors = collect_rows(
        base_url,
        token,
        args.group_id,
        args.days,
        args.max_projects,
        args.delay,
    )
    write_excel(args.excel.expanduser(), rows, errors)
    print(f"Wrote {args.excel.expanduser()} ({len(rows)} projects)")
    if errors:
        print(
            f"{len(errors)} projects had API errors; see the Errors sheet.",
            file=sys.stderr,
        )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
